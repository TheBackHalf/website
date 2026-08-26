from __future__ import annotations

import json
import shutil
import tempfile
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

WORKBOOK = Path(
    r"C:\Users\smyle\OneDrive\Desktop\The Back Half\The Back Half\Founder Command Center\The_Back_Half_Founder_Command_Center_8.25.2026 through Launch.xlsx"
)
OUT = Path(__file__).resolve().parents[2] / "ops" / "fab-5" / "aos-command-center-snapshot.json"


def copy_readable(src: Path) -> Path:
    tmp = Path(tempfile.gettempdir()) / "fcc_aos_readonly.xlsx"
    shutil.copy2(src, tmp)
    return tmp


def rows_from_sheet(ws, tab: str) -> list[dict]:
    items = []
    for excel_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        values = list(row)
        if not values:
            continue
        order = values[0]
        if order is None and not (len(values) > 2 and values[2]):
            continue
        if not isinstance(order, (int, float)):
            continue
        if tab == "August Launch":
            deliverable = values[2] if len(values) > 2 else ""
            description = values[3] if len(values) > 3 else ""
            percent = values[4] if len(values) > 4 else 0
            status = values[5] if len(values) > 5 else ""
            dependency = values[6] if len(values) > 6 else ""
            cost = values[7] if len(values) > 7 else ""
            assigned = values[8] if len(values) > 8 else ""
            notes = values[9] if len(values) > 9 else ""
            timing = ""
        else:
            deliverable = values[2] if len(values) > 2 else ""
            description = values[3] if len(values) > 3 else ""
            percent = values[4] if len(values) > 4 else 0
            status = values[5] if len(values) > 5 else ""
            dependency = values[6] if len(values) > 6 else ""
            cost = values[7] if len(values) > 7 else ""
            timing = values[8] if len(values) > 8 else ""
            assigned = values[9] if len(values) > 9 else ""
            notes = values[3] if len(values) > 3 else ""
        items.append(
            {
                "tab": tab,
                "excelRow": excel_row,
                "order": int(order) if isinstance(order, (int, float)) else order,
                "phase": str(values[1] or "") if len(values) > 1 else "",
                "deliverable": str(deliverable or ""),
                "description": str(description or ""),
                "percentComplete": float(percent or 0),
                "status": str(status or ""),
                "dependency": str(dependency or ""),
                "cost": str(cost or ""),
                "assignedAgent": str(assigned or ""),
                "notes": str(notes or ""),
                "targetTiming": str(timing or ""),
            }
        )
    return items


def main() -> None:
    readable = copy_readable(WORKBOOK)
    wb = load_workbook(readable, read_only=True, data_only=True)
    august = rows_from_sheet(wb["August Launch"], "August Launch")
    post = rows_from_sheet(wb["Post Launch"], "Post Launch")
    payload = {
        "ingestedAt": datetime.now(timezone.utc).isoformat(),
        "workbook": str(WORKBOOK),
        "authoritativeTabPreLaunch": "August Launch",
        "authoritativeTabPostLaunch": "Post Launch",
        "rows": august + post,
    }
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf8")
    print(f"WROTE {OUT} rows={len(payload['rows'])}")


if __name__ == "__main__":
    main()
