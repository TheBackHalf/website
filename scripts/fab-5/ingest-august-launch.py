"""Ingest the authoritative August Launch tab into ops/fab-5/launch-rows.json."""

from __future__ import annotations

import json
import re
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

AUTHORITATIVE_WORKBOOK = Path(
    r"c:\Users\smyle\OneDrive\Desktop\The Back Half\The Back Half\Founder Command Center\The_Back_Half_Founder_Command_Center_8.16.2026 through Launch.xlsx"
)
TAB = "August Launch"
REPO = Path(__file__).resolve().parents[2]
OUT = REPO / "ops" / "fab-5" / "launch-rows.json"
READABLE_COPY = REPO / ".tmp-launch-ingest" / "command-center-8.16.2026-through-launch.xlsx"
VENDOR_RE = re.compile(
    r"bank|stripe|insurance|trademark|vendor|hosting|domain|dns|ssl|counsel|broker|tax professional|filing",
    re.I,
)

# workstream, owner, supporting, founderAction, humanExpert, nextAction
R: dict[int, tuple] = {}


def a(order, ws, owner, supporting, founder, expert, nxt):
    R[order] = (ws, owner, supporting, founder, expert, nxt)


# Remaining / in-progress / planned assignments (workbook statuses preserved except Row 18).
a(6, "INNOVATION", "nia", ["michelle"], False, False, "Create the Version 1 deferred-enhancement register from already-identified non-launch items. Do not add launch scope.")
a(10, "FINANCE", "kimberly", ["michelle"], True, False, "Open operating, tax-reserve, and payment-settlement accounts. Founder identity/signature required at the bank.")
a(11, "FINANCE", "kimberly", ["michelle"], False, False, "Stand up chart of accounts, expense categories, revenue tracking, and receipt process after banking exists.")
a(12, "FINANCE", "kimberly", ["michelle"], True, False, "Confirm remaining licenses/registrations. Founder signs government filings where required.")
a(13, "FINANCE", "kimberly", ["michelle"], True, True, "Establish the tax filing calendar. Qualified tax professional for tax conclusions; Founder for filings.")
a(14, "FINANCE", "kimberly", ["imani", "michelle"], True, True, "Determine checkout tax treatment with qualified tax professional; Imani configures Stripe only after the decision.")
a(18, "OPERATIONS", "michelle", ["imani", "nia", "kimberly"], False, False, "Founder review of Row 18 workstream assignments. Do not start Row 19.")
a(19, "OPERATIONS", "michelle", ["imani", "nia"], False, False, "NONE — Founder accepted. Do not start Row 20.")
a(20, "TECHNOLOGY", "imani", ["michelle"], True, False, "FOUNDER ACTION REQUIRED: authorize scoped Vercel deploy/log access for Imani and delegated support@/privacy@ mailbox access for Michelle. Do not paste secrets in chat. Do not start Row 21.")
a(21, "OPERATIONS", "michelle", ["kimberly"], False, False, "Separate Founder-only tasks into one queue from the full launch view. Do not send routine work to Kimberly.")
a(23, "LEGAL", "imani", ["michelle", "kimberly"], True, True, "File approved launch-critical trademarks. IP counsel for filings; Founder signature. Imani does not give legal conclusions.")
a(32, "LEGAL", "imani", ["nia", "michelle"], False, False, "Audit that approved legal documents are published, linked, consented, and versioned in production. Implement only; no legal rewrite.")
a(33, "MARKETING", "nia", ["imani", "michelle"], False, False, "Write the marketing-claims and testimonial standard from approved legal positioning. Escalate legally sensitive claims.")
a(34, "LEGAL", "imani", ["michelle", "nia"], False, True, "Escalate launch-critical legal documents and consent flows to qualified human counsel. Agents do not conclude the law.")
a(49, "EXPERIENCE", "nia", ["michelle"], False, False, "Create Spanish Founder video scripts from the Spanish Voice Bible and approved English scripts. Do not invent curriculum.")
a(50, "EXPERIENCE", "nia", ["imani", "michelle"], False, False, "Finish Founder video production, captions, hosting, and acceptance from approved English/Spanish scripts and AI Kimberly assets.")
a(51, "LEARNING", "nia", ["imani"], False, False, "Finish printable Blueprint/journal assets from already-approved guidebook content.")
a(60, "LEGAL", "imani", ["michelle", "nia", "kimberly"], True, True, "Founder decides minimum participant age; Imani implements enforcement; human legal review if the rule is a legal conclusion.")
a(61, "TECHNOLOGY", "imani", ["michelle"], False, False, "Configure production uptime, error, database, and payment monitoring.")
a(62, "TECHNOLOGY", "imani", ["michelle"], False, False, "Prove automated backups with a successful restore test.")
a(72, "TECHNOLOGY", "imani", ["michelle", "kimberly"], False, False, "Document every launch-critical vendor, owner, billing, renewal, limits, and fallback. Do not invent vendors.")
a(73, "TECHNOLOGY", "imani", ["kimberly", "michelle"], False, False, "Verify production plans, payment methods, and quotas cannot halt registration, Lumina, email, hosting, or payments.")
a(74, "TECHNOLOGY", "imani", ["michelle", "kimberly"], True, False, "Build the credential/MFA recovery register. Founder for identity/MFA on Founder-held accounts. Never store secrets in the register values.")
a(75, "TECHNOLOGY", "imani", ["michelle"], False, False, "Validate domain, DNS, SSL, auto-renewal, and recovery access.")
a(76, "MARKETING", "nia", ["imani", "michelle"], False, False, "Stand up official social channels using approved brand assets. Do not add unapproved networks.")
a(77, "MARKETING", "nia", ["michelle", "imani"], False, False, "Write social channel governance, access, MFA, and posting authority.")
a(78, "MARKETING", "nia", ["michelle"], False, False, "Build the 7-day pre-launch campaign from approved assets.")
a(79, "MARKETING", "nia", ["michelle"], False, False, "Build launch-day social content against final production URLs.")
a(80, "MARKETING", "nia", ["michelle"], False, False, "Prepare the first 30-day post-launch content plan. Do not expand product scope.")
a(81, "MARKETING", "nia", ["imani"], False, False, "Produce platform-native social assets from approved creative.")
a(82, "MARKETING", "nia", ["imani", "michelle"], False, False, "Implement social publishing/scheduling once channels are named and connected.")
a(83, "MARKETING", "nia", ["michelle"], False, False, "Implement social engagement/response protocol with support routing.")
a(84, "MARKETING", "nia", ["imani", "michelle"], False, False, "Create the launch marketing KPI dashboard after analytics events exist.")
a(85, "MARKETING", "nia", ["imani"], False, False, "Validate social → site → checkout → purchase attribution.")
a(86, "MARKETING", "nia", ["michelle"], False, False, "Finalize audience/partner outreach lists from approved communications.")
a(87, "MARKETING", "nia", ["michelle"], False, False, "Execute the pre-launch campaign only after scheduling and assets are ready.")
a(88, "MARKETING", "nia", ["michelle"], False, False, "Complete public brand-presence audit across web/social.")
a(121, "EXPERIENCE", "nia", ["imani"], False, False, "Finish AI Kimberly production assets from already-approved Codex, wardrobe, and Kim’s Space materials.")
a(122, "TECHNOLOGY", "imani", ["kimberly", "michelle"], False, False, "Set production AI usage, spend alerts, rate limits, timeouts, and degraded-experience fallbacks. Kimberly only for material unbudgeted spend.")
a(133, "LEARNING", "nia", ["imani"], False, False, "Specify progression/save rules; Imani implements pause/resume, drafts, and loss-prevention. Nia verifies teachability.")
a(134, "LEARNING", "nia", ["imani"], False, False, "Specify Back Half Portfolio contents from approved artifacts; Imani implements assembly/download.")
a(135, "LEARNING", "nia", ["imani", "michelle"], False, False, "Specify completion/threshold ceremony; Imani implements; Nia verifies Triple E.")
a(136, "EXPERIENCE", "nia", ["imani"], False, False, "Audit downloadable Blueprint/participant PDFs against the accessibility standard.")
a(137, "EXPERIENCE", "nia", ["imani"], False, False, "Produce and QA Spanish Founder videos from approved Spanish scripts and Voice Bible.")
a(138, "EXPERIENCE", "nia", ["imani"], False, False, "Add captions/transcripts/posters for every launch-critical Founder video.")
a(139, "LEARNING", "nia", ["imani"], False, False, "Produce Spanish Blueprint and downloadable participant assets from approved translation + human Spanish QA.")
a(140, "LEGAL", "imani", ["nia", "michelle"], False, True, "Publish full approved English and Spanish legal bodies. Human legal review if equivalence is a legal-meaning question. Nia verifies presentation.")
a(141, "LEGAL", "imani", ["michelle"], False, True, "Inventory subprocessors and complete third-party/AI privacy review with qualified human counsel for legal conclusions.")
a(142, "EXPERIENCE", "nia", ["imani"], False, False, "QA bilingual AI Kimberly so Español mode uses approved Spanish UI and Founder voice.")
a(145, "EXPERIENCE", "nia", ["imani"], False, False, "Build bilingual email templates from the approved English email library and Spanish QA.")
a(146, "TECHNOLOGY", "imani", ["michelle"], False, False, "Configure transactional email: domain, provider, suppression, bounce, unsubscribe, monitoring.")
a(147, "TECHNOLOGY", "imani", ["nia", "michelle"], False, False, "Build lifecycle automations from the approved automation map and platform events.")
a(148, "TECHNOLOGY", "imani", ["nia"], False, False, "Test email deliverability across major inboxes, mobile, links, and legal footer.")
a(150, "TECHNOLOGY", "imani", ["michelle"], False, False, "Instrument launch-critical analytics events across site, checkout, Journey, and Lumina.")
a(151, "OPERATIONS", "michelle", ["imani"], False, False, "Build the daily launch dashboard from event tracking.")
a(153, "OPERATIONS", "michelle", ["imani", "nia"], False, False, "Connect support@ intake, acknowledgment, categories, and urgent escalation.")
a(155, "OPERATIONS", "michelle", ["imani", "nia"], False, False, "Operationalize Architect support: inbox, ownership, SLA, payment/access escalation.")
a(156, "OPERATIONS", "michelle", ["nia"], False, False, "Publish launch support knowledge-base answers for account, payment, Journey, Lumina, and cancellation.")
a(157, "OPERATIONS", "michelle", ["imani", "nia"], False, False, "Assign first-72-hour support coverage and severity paths into the runbook.")
a(158, "EXPERIENCE", "nia", ["michelle"], False, False, "Create the Voice-of-Architect capture method for launch-day feedback.")
a(159, "COMMUNITY", "nia", ["michelle", "imani"], False, False, "Finalize Founding Architect Community runway messaging from locked pricing/community decisions. Do not invent a new community program.")
a(160, "COMMUNITY", "nia", ["michelle"], False, False, "Create Founding Architect welcome/retention plan from purchase through Community opening.")
a(161, "TECHNOLOGY", "imani", ["michelle"], False, False, "Validate business email deliverability for all production senders.")
a(162, "LEGAL", "imani", ["nia", "michelle"], False, False, "Implement marketing-email compliance and suppression. Human expert only if a new legal conclusion is required.")
a(163, "FINANCE", "kimberly", ["imani", "michelle"], False, False, "Operationalize daily Stripe/bank/entitlement reconciliation after banking and accounting exist.")
a(164, "FINANCE", "kimberly", ["michelle"], False, False, "Create remaining launch cash/spend control against the $5,000 cap. No unplanned vendor spend.")
a(165, "FINANCE", "kimberly", ["imani", "nia"], False, False, "Independently verify displayed price, checkout, tax, receipts, and failed-payment behavior against locked pricing.")
a(166, "FINANCE", "kimberly", ["michelle"], True, False, "Assess and bind launch insurance. Founder binds coverage; this is an external broker/vendor step, not agent legal advice.")
a(167, "LEGAL", "imani", ["michelle"], False, False, "Operationalize privacy-rights requests: access, deletion, export, identity verification, tracking.")
a(168, "LEGAL", "imani", ["nia", "michelle"], False, False, "Validate cookie/tracking consent against actual production tags and the approved Privacy Policy.")
a(169, "LEGAL", "imani", ["michelle"], False, True, "Extend incident response to privacy/security breach notification. Human legal for notification-law conclusions.")
a(170, "FINANCE", "kimberly", ["imani", "michelle", "nia"], False, False, "Operationalize chargebacks, disputes, fraud, and payment failures with Stripe + support routing.")
a(171, "TECHNOLOGY", "imani", ["kimberly", "nia", "michelle"], False, False, "Validate cancellation, renewal, and access lifecycle against locked entitlements and Community rules.")
a(172, "FINANCE", "kimberly", ["michelle"], False, False, "Create the corporate compliance/records calendar from entity, tax, insurance, and IP dates.")
a(173, "FINANCE", "kimberly", ["michelle"], False, False, "Establish monthly bookkeeping and financial close after accounting exists.")
a(174, "OPERATIONS", "michelle", ["imani", "kimberly"], True, False, "Document Founder-unavailable continuity using the accepted Fab 5 operating system. Do not expand agent authority.")
a(176, "TECHNOLOGY", "imani", ["michelle", "nia"], False, False, "Create English/Spanish test accounts and data from the master test plan.")
a(177, "TECHNOLOGY", "imani", ["nia", "michelle"], False, False, "Run functional testing of the feature-complete build. Nia verifies participant-facing failures.")
a(178, "TECHNOLOGY", "imani", ["nia", "michelle"], False, False, "Run discovery-to-completion enrollment E2E.")
a(179, "TECHNOLOGY", "imani", ["kimberly", "michelle"], False, False, "Run payment and subscription testing including failure, installment, and webhook recovery.")
a(180, "EXPERIENCE", "nia", ["imani"], False, False, "Run the Lumina evaluation suite. Imani supplies runtime evidence; Nia accepts voice/safety/bilingual quality.")
a(181, "TECHNOLOGY", "imani", ["michelle"], False, False, "Run data integrity and recovery testing.")
a(182, "TECHNOLOGY", "imani", ["michelle"], False, False, "Run security review against the security baseline.")
a(183, "EXPERIENCE", "nia", ["imani"], False, False, "Run accessibility review. Imani implements fixes; Nia accepts.")
a(184, "TECHNOLOGY", "imani", ["nia"], False, False, "Run cross-browser compatibility testing of launch-critical flows.")
a(185, "TECHNOLOGY", "imani", ["michelle"], False, False, "Run performance, capacity, and load testing.")
a(186, "TECHNOLOGY", "imani", ["nia"], False, False, "Test priority iPhone/Android devices.")
a(187, "TECHNOLOGY", "imani", ["nia"], False, False, "Fix responsive/touch defects found in mobile testing. Nia retests participant experience.")
a(188, "OPERATIONS", "michelle", ["imani", "nia"], False, False, "Triage all defects with owner, severity, and release requirement.")
a(189, "TECHNOLOGY", "imani", ["nia", "michelle"], False, False, "Resolve critical/high defects. Nia retests participant-facing fixes.")
a(190, "TECHNOLOGY", "imani", ["michelle", "nia"], False, False, "Run regression after critical fixes.")
a(191, "OPERATIONS", "michelle", ["nia", "imani"], False, False, "Run the private founding experience under controlled conditions. Not a public launch.")
a(192, "EXPERIENCE", "nia", ["michelle"], False, False, "Collect structured validation results from the private founding experience.")
a(193, "TECHNOLOGY", "imani", ["nia", "michelle"], False, False, "Resolve validation blockers. Defer only documented noncritical enhancements into row 6.")
a(194, "TECHNOLOGY", "imani", ["michelle"], False, False, "Complete production security audit.")
a(195, "TECHNOLOGY", "imani", ["michelle"], False, False, "Validate analytics events and dashboards before launch.")
a(196, "LEARNING", "nia", ["imani"], False, False, "Complete Journey instructional-integrity audit of Chapters I–VII against approved manuscript structure.")
a(197, "LEARNING", "nia", ["imani"], False, False, "Validate Blueprint-to-Journey alignment and saved-answer population.")
a(198, "EXPERIENCE", "nia", ["michelle"], False, False, "Nia independently Triple-E reviews the actual production experience. Self-report is not acceptance.")
a(199, "MARKETING", "nia", ["michelle"], False, False, "Finish launch communications: founder video, email, LinkedIn/Instagram, FAQs, support scripts.")
a(200, "MARKETING", "nia", ["michelle", "imani"], False, False, "Schedule and link-test all approved August 19 communications.")
a(201, "OPERATIONS", "michelle", ["imani", "nia", "kimberly"], True, False, "Confirm launch-day and 72-hour coverage. Founder only for Founder-coverage hours, not routine staffing.")
a(202, "OPERATIONS", "michelle", ["imani", "nia"], False, False, "Finish the launch-day runbook: timed release, monitoring, comms, support, rollback, authority.")
a(203, "OPERATIONS", "michelle", ["imani"], False, False, "Create production change-control from final QA through stabilization.")
a(204, "OPERATIONS", "michelle", ["imani", "nia"], False, False, "Create the launch evidence repository. A status cell is not evidence.")
a(205, "OPERATIONS", "michelle", ["imani", "nia"], False, False, "Re-audit previously Complete rows against acceptance evidence, not file existence.")
a(206, "OPERATIONS", "michelle", ["imani", "nia"], False, False, "Execute the full go-live simulation from the regression-passed candidate and runbook.")
a(208, "OPERATIONS", "michelle", ["imani", "nia"], False, False, "Establish the hypercare command center for the first 30 days.")
a(209, "OPERATIONS", "michelle", ["imani"], False, False, "Create the launch-day executive dashboard from analytics and monitoring.")
a(210, "MARKETING", "nia", ["michelle"], False, False, "Create the launch-day communications command plan: who publishes, verifies, pauses.")
a(211, "OPERATIONS", "michelle", ["nia", "imani", "kimberly"], False, False, "Create the 7-day executive post-launch plan.")
a(212, "OPERATIONS", "michelle", ["kimberly", "nia"], False, False, "Create the 30-day launch performance review framework.")
a(213, "OPERATIONS", "michelle", ["imani", "nia", "kimberly"], False, False, "Run Fab 5 functional Go/No-Go reviews with evidence per workstream. Michelle coordinates; specialists certify their domains.")
a(214, "EXPERIENCE", "nia", ["imani", "michelle"], False, False, "Independently test public discovery-to-purchase as a stranger.")
a(215, "OPERATIONS", "michelle", ["imani", "nia"], False, False, "Verify Founder-absence operating capability using the accepted Fab 5 model.")
a(216, "OPERATIONS", "michelle", ["imani", "nia", "kimberly"], False, False, "Complete the final business-ready acceptance audit across legal, finance, product, and operations.")
a(217, "FINANCE", "kimberly", ["michelle", "imani", "nia"], True, False, "Founder Go/No-Go. Michelle packages evidence. Do not self-certify launch.")
a(218, "TECHNOLOGY", "imani", ["michelle", "kimberly"], True, False, "Freeze production after Go. Founder for release-despite-blocker; Imani executes freeze.")
a(219, "TECHNOLOGY", "imani", ["michelle", "nia"], False, False, "Publish website and open enrollment on August 19 after freeze.")
a(220, "TECHNOLOGY", "imani", ["michelle", "nia"], False, False, "Activate production platform: registration, payments, Lumina, Journey, email, analytics, support.")
a(221, "MARKETING", "nia", ["michelle", "kimberly"], True, False, "Release Founder announcement from approved assets. Founder identity/approval for the announcement.")
a(222, "OPERATIONS", "michelle", ["imani", "nia"], False, False, "Monitor first-day operations from the executive dashboard.")
a(223, "TECHNOLOGY", "imani", ["michelle", "nia"], False, False, "Resolve launch-day incidents per the incident plan.")
a(224, "OPERATIONS", "michelle", ["imani", "nia", "kimberly"], False, False, "Complete end-of-day executive review.")
a(225, "OPERATIONS", "michelle", ["imani", "nia"], False, False, "Begin 72-hour stabilization.")

CRITICAL_PATH = [
    18, 19, 20, 10, 11, 133, 134, 135, 32, 34, 50, 121,
    176, 177, 178, 179, 188, 189, 190, 206, 213, 216, 217, 218, 219, 220,
]

DUPLICATES = [
    {
        "rowA": 19,
        "rowB": 15,
        "overlap": "Executive authority/escalation is already encoded in the Founder-accepted Row 15 operating system.",
        "distinctAcceptance": "Row 19 is a separate August Launch deliverable. Founder accepted it as 100% complete. It operationalizes the accepted Row 15–18 authority model and does not replace it.",
        "recommendation": "Keep Row 19. Treat Row 15 OS as the source to publish, not a new operating model. Do not delete.",
    },
    {
        "rowA": 55,
        "rowB": 156,
        "overlap": "Support knowledge-base content vs launch operational KB.",
        "distinctAcceptance": "Row 55 finalized Help Center content; Row 156 operationalizes launch-time answers and routing.",
        "recommendation": "Keep both. Distinct acceptance: content approval vs operating KB.",
    },
    {
        "rowA": 143,
        "rowB": 147,
        "overlap": "Lifecycle automation map vs build.",
        "distinctAcceptance": "Row 143 mapped triggers; Row 147 implements them.",
        "recommendation": "Keep both.",
    },
    {
        "rowA": 152,
        "rowB": 155,
        "overlap": "Support workflow design vs operating support function.",
        "distinctAcceptance": "Row 152 created the workflow; Row 155 operationalizes inbox/SLA/ownership.",
        "recommendation": "Keep both. Also related to Row 153 channel configuration.",
    },
    {
        "rowA": 32,
        "rowB": 34,
        "overlap": "Legal implementation audit vs human legal review.",
        "distinctAcceptance": "Row 32 is publication/consent implementation evidence; Row 34 is qualified human counsel judgment.",
        "recommendation": "Keep both. Do not treat Imani implementation as legal conclusions.",
    },
    {
        "rowA": 6,
        "rowB": None,
        "overlap": "Version 1 Deferred-Enhancement Register already exists as August Launch Row 6. Future Expansion tab also exists in the workbook.",
        "distinctAcceptance": "Row 6 is the launch-plan register; Future Expansion is a separate workbook tab.",
        "recommendation": "Do not create another deferred-enhancement tracker. Use Row 6.",
    },
]


def classify_complete(order: int, phase: str, title: str) -> tuple:
    t = title.lower()
    if order in (1, 2):
        return "OPERATIONS", "kimberly", ["michelle"], True, False
    if order == 3:
        return "FINANCE", "kimberly", ["michelle"], True, False
    if order in (15, 17):
        return "OPERATIONS", "michelle", ["imani", "nia"], False, False
    if order == 16:
        return "TECHNOLOGY", "imani", ["michelle", "nia"], False, False
    if any(k in t for k in ("trademark", "terms", "privacy", "participant agreement", "membership agreement", "ai disclosure", "consent", "ip ownership", "copyright")):
        return "LEGAL", "imani", ["michelle", "nia"], False, False
    if "community guidelines" in t:
        return "COMMUNITY", "nia", ["michelle", "imani"], False, False
    if any(k in t for k in ("chapter", "journey manuscript", "aliveness assessment", "journey exercises", "blueprint", "journal")) and "page" not in t:
        return "LEARNING", "nia", ["imani"], False, False
    if any(k in t for k in ("lumina", "founder video", "founder story", "voice bible", "founder media", "ai kimberly", "welcome")):
        return "EXPERIENCE", "nia", ["imani"], False, False
    if any(k in t for k in ("social", "email templates", "category narrative", "lexicon", "brand style", "logo", "typography", "visual ip", "image library", "website copy")):
        return "MARKETING" if any(k in t for k in ("social", "email templates", "category")) else "EXPERIENCE", "nia", ["michelle"], False, False
    if any(k in t for k in ("stripe", "checkout", "webhook", "billing", "hosting", "database", "auth", "login", "registration", "password", "role and access", "application shell", "seo", "domain", "security baseline", "data model", "system architecture", "build stack", "environments")):
        return "TECHNOLOGY", "imani", ["michelle"], False, False
    if any(k in t for k in ("ein", "legal entity", "llc", "banking", "budget")):
        return "FINANCE", "kimberly", ["michelle"], False, False
    if "support workflow" in t or "incident response" in t or "operating rhythm" in t or "decision log" in t or "sops" in t or "rollback" in t or "test plan" in t:
        owner = "imani" if "rollback" in t else "michelle"
        ws = "TECHNOLOGY" if owner == "imani" else "OPERATIONS"
        return ws, owner, ["imani", "nia"] if owner == "michelle" else ["michelle"], False, False
    if phase.startswith("Business"):
        return "FINANCE", "kimberly", ["michelle"], False, False
    if phase.startswith("Technology"):
        return "TECHNOLOGY", "imani", ["michelle"], False, False
    if phase.startswith("Parallel Marketing"):
        return "MARKETING", "nia", ["michelle"], False, False
    if phase.startswith("Website"):
        return "TECHNOLOGY", "imani", ["nia", "michelle"], False, False
    if phase.startswith("Lumina"):
        return "EXPERIENCE", "nia", ["imani"], False, False
    if phase.startswith("Journey"):
        return "LEARNING", "nia", ["imani"], False, False
    if phase.startswith("Email"):
        return "OPERATIONS", "michelle", ["imani", "nia"], False, False
    if phase.startswith("Testing"):
        return "OPERATIONS", "michelle", ["imani", "nia"], False, False
    if phase.startswith("Legal"):
        return "EXPERIENCE", "nia", ["imani"], False, False
    if phase.startswith("Governance"):
        return "OPERATIONS", "michelle", ["kimberly"], False, False
    if phase.startswith("Final"):
        return "OPERATIONS", "michelle", ["imani", "nia"], False, False
    return "OPERATIONS", "michelle", ["imani", "nia"], False, False


def workbook_path() -> Path:
    for path in (AUTHORITATIVE_WORKBOOK, READABLE_COPY):
        if path.exists():
            try:
                path.open("rb").close()
                return path
            except OSError:
                continue
    raise FileNotFoundError("August Launch workbook not readable from known Founder Command Center paths.")


def apply_dependency_maps(records: list[dict], remaining: list[dict]) -> None:
    rem_by_id = {item["spreadsheetRow"]: item for item in remaining}
    unlocks: dict[int, list[int]] = {item["spreadsheetRow"]: [] for item in remaining}
    for item in remaining:
        dep = (item.get("originalDependency") or "").lower()
        for other in remaining:
            title = (other.get("deliverable") or "").lower()
            if len(title) >= 12 and title in dep and other["spreadsheetRow"] != item["spreadsheetRow"]:
                unlocks[other["spreadsheetRow"]].append(item["spreadsheetRow"])
    for item in remaining:
        number = item["spreadsheetRow"]
        dep = item.get("originalDependency") or "-"
        blocked_by_remaining = []
        for other in remaining:
            title = (other.get("deliverable") or "").lower()
            if len(title) >= 12 and title in dep.lower() and other["spreadsheetRow"] != number:
                blocked_by_remaining.append(other["spreadsheetRow"])
        parallel = []
        my_title = (item.get("deliverable") or "").lower()
        for other in remaining:
            if other["spreadsheetRow"] == number or other["primaryOwner"] == item["primaryOwner"]:
                continue
            other_dep = (other.get("originalDependency") or "").lower()
            if len(my_title) >= 12 and my_title in other_dep:
                continue
            if other["spreadsheetRow"] in blocked_by_remaining:
                continue
            if abs(other["spreadsheetRow"] - number) <= 40:
                parallel.append(other["spreadsheetRow"])
            if len(parallel) >= 6:
                break
        mapping = {
            "blockedBy": dep if dep not in ("-", "", None) else "none",
            "unlocks": unlocks.get(number, []),
            "canRunInParallelWith": parallel,
            "requiresFounder": item["founderActionRequired"],
            "requiresHumanExpert": item["humanExpertRequired"],
            "requiresExternalVendorService": bool(
                VENDOR_RE.search(f"{item.get('deliverable') or ''} {item.get('description') or ''} {dep}")
            ),
        }
        item["dependency"] = dep
        item["crossFunctionalDependencies"] = mapping
        item["unlocks"] = mapping["unlocks"]
        item["canRunInParallelWith"] = mapping["canRunInParallelWith"]
        item["requiresExternalVendorService"] = mapping["requiresExternalVendorService"]
    for rec in records:
        mapped = rem_by_id.get(rec["number"])
        rec["crossFunctionalDependencies"] = (
            mapped["crossFunctionalDependencies"]
            if mapped
            else {
                "blockedBy": "none — complete",
                "unlocks": [],
                "canRunInParallelWith": [],
                "requiresFounder": rec["founderActionRequired"],
                "requiresHumanExpert": rec["humanExpertRequired"],
                "requiresExternalVendorService": False,
            }
        )


def load_rows(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[TAB]
    rows = []
    for excel_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        values = list(row) + [None] * 8
        order, phase, deliverable, desc, pct, status, dep, cost = values[:8]
        if order is None and not deliverable:
            continue
        rows.append(
            {
                "excelRow": excel_row,
                "order": int(order) if isinstance(order, (int, float)) else order,
                "phase": phase,
                "deliverable": deliverable,
                "description": desc,
                "percentComplete": pct,
                "status": status,
                "dependency": "-" if dep is None else str(dep),
                "cost": None if cost is None else str(cost),
            }
        )
    return rows


def main() -> None:
    source = workbook_path()
    raw = load_rows(source)
    if len(raw) < 200:
        raise SystemExit(f"Expected full August Launch tab, found {len(raw)} rows")

    records = []
    remaining = []
    for item in raw:
        order = item["order"]
        status = str(item["status"] or "")
        complete = status.lower() == "complete"
        if order in R:
            ws, owner, supporting, founder, expert, nxt = R[order]
        else:
            ws, owner, supporting, founder, expert = classify_complete(order, item["phase"] or "", item["deliverable"] or "")
            nxt = "NONE — Complete in August Launch tab"
        acceptance = "complete" if complete else "open"
        founder_acc = None
        evidence = []
        pct = item["percentComplete"]
        out_status = status
        if order == 15:
            founder_acc = "accepted"
            acceptance = "founder_accepted"
            evidence = ["ops/fab-5/operating-system.json", "ops/fab-5/acceptance-record.json"]
        elif order == 16:
            founder_acc = "accepted"
            acceptance = "founder_accepted"
            evidence = ["lib/fab-5/", "ops/fab-5/acceptance-record.json", "ops/fab-5/runs/row-16-live-smoke.json"]
        elif order == 17:
            founder_acc = "accepted"
            acceptance = "founder_accepted"
            evidence = ["ops/fab-5/runs/row-17-autonomy-validation.json", "ops/fab-5/acceptance-record.json"]
        elif order == 18:
            out_status = "Complete"
            pct = 100
            founder_acc = "accepted"
            acceptance = "founder_accepted"
            nxt = "NONE — Founder accepted. Do not start Row 19."
            evidence = [
                "ops/fab-5/launch-rows.json",
                "lib/fab-5/workstreams.ts",
                "ops/fab-5/runs/row-18-launch-workstreams.json",
                "ops/fab-5/acceptance-record.json",
            ]
            complete = True
        elif order == 19:
            out_status = "Complete"
            pct = 100
            founder_acc = "accepted"
            acceptance = "founder_accepted"
            nxt = "NONE — Founder accepted. Do not start Row 20."
            founder = False
            evidence = [
                "ops/fab-5/operating-system.json",
                "lib/fab-5/authority.ts",
                "ops/fab-5/runs/row-19-authority-escalation-validation.json",
                "ops/fab-5/acceptance-record.json",
            ]
            complete = True
        rec = {
            "number": order,
            "excelRow": item["excelRow"],
            "phase": item["phase"],
            "deliverable": item["deliverable"],
            "description": item["description"],
            "percentComplete": pct,
            "status": out_status,
            "workbookStatus": status,
            "dependency": item["dependency"],
            "dependencies": [] if item["dependency"] in ("-", "", None) else [item["dependency"]],
            "cost": item["cost"],
            "primaryWorkstream": ws,
            "primaryOwner": owner,
            "supportingOwners": supporting,
            "founderActionRequired": founder,
            "humanExpertRequired": expert,
            "nextAction": nxt,
            "evidenceRequired": "Workbook acceptance evidence plus production/runtime proof where the deliverable is operational.",
            "criticalPathState": "on_path" if order in CRITICAL_PATH else ("complete" if complete else "parallel_or_later"),
            "acceptanceState": acceptance,
            "founderAcceptance": founder_acc,
            "remainingLaunchCritical": not complete,
            "mutating": False,
            "evidence": evidence,
            "blockers": [] if complete else ([item["dependency"]] if item["dependency"] not in ("-", "", None) else []),
        }
        records.append(rec)
        if not complete:
            remaining.append(
                {
                    "id": str(order),
                    "spreadsheetRow": order,
                    "phase": rec["phase"],
                    "deliverable": rec["deliverable"],
                    "description": rec["description"],
                    "percentComplete": rec["percentComplete"],
                    "status": rec["status"],
                    "primaryWorkstream": ws,
                    "primaryOwner": owner,
                    "supportingOwners": supporting,
                    "dependencies": rec["dependencies"],
                    "originalDependency": item["dependency"],
                    "blockers": rec["blockers"],
                    "founderActionRequired": founder,
                    "humanExpertRequired": expert,
                    "evidenceRequiredForCompletion": rec["evidenceRequired"],
                    "cost": rec["cost"],
                    "nextAction": nxt,
                    "evidenceAcceptanceState": acceptance,
                    "priority": order,
                    "criticalPath": order in CRITICAL_PATH,
                    "blockedBy": item["dependency"],
                    "requiresFounder": founder,
                    "requiresHumanExpert": expert,
                    "source": [str(AUTHORITATIVE_WORKBOOK), f"{TAB} row {order}"],
                }
            )

    apply_dependency_maps(records, remaining)
    owners = Counter(r["primaryOwner"] for r in remaining)
    statuses = Counter(r["workbookStatus"] for r in records)
    unowned = [r for r in remaining if r["primaryOwner"] not in {"kimberly", "michelle", "imani", "nia"}]
    if unowned:
        raise SystemExit(f"Unowned remaining: {unowned}")

    used_classifier_fallback = [
        r["spreadsheetRow"] for r in remaining if r["spreadsheetRow"] not in R
    ]
    adapter = {
        "adapter": "Machine-readable representation of the authoritative August Launch tab. Does not replace the workbook.",
        "authoritativeWorkbook": str(AUTHORITATIVE_WORKBOOK),
        "ingestedFromReadableCopy": None if source == AUTHORITATIVE_WORKBOOK else str(source),
        "authoritativeTab": TAB,
        "ingestedAt": "2026-08-17",
        "sourceLimitation": None if source == AUTHORITATIVE_WORKBOOK else "Original workbook was locked; ingested from a readable copy of the same file.",
        "reconciliationExceptions": [
            {
                "item": f"Remaining row {n} was not in the explicit remaining-assignment table; classified by phase/title.",
                "whyFlagged": "Ownership still assigned; verify if a new remaining row appeared.",
            }
            for n in used_classifier_fallback
        ],
        "row18": {
            "number": 18,
            "deliverable": "Create Fab 5 Launch Workstreams",
            "technicalStatus": "complete",
            "founderAcceptance": "accepted",
            "founderAcceptedAt": "2026-08-17",
            "row19Started": True,
        },
        "row19": {
            "number": 19,
            "deliverable": "Establish Executive Authority and Escalation Matrix",
            "technicalStatus": "complete",
            "founderAcceptance": "accepted",
            "founderAcceptedAt": "2026-08-17",
            "row20Started": False,
        },
        "statusCounts": {
            "total": len(records),
            "complete": sum(1 for r in records if str(r["status"]).lower() == "complete"),
            "inProgress": statuses.get("In Progress", 0),
            "notStarted": sum(
                1
                for r in records
                if str(r["workbookStatus"]) == "Not Started" and str(r["status"]).lower() != "complete"
            ),
            "planned": statuses.get("Planned", 0),
            "blockedOther": 0,
            "remaining": len(remaining),
            "remainingWithPrimaryOwner": len(remaining),
            "unownedRemaining": 0,
            "row18AcceptedSnapshot": {
                "total": 225,
                "remainingAtAcceptance": 125,
                "remainingWithPrimaryOwnerAtAcceptance": 125,
                "unownedRemaining": 0,
            },
            "row19AcceptedSnapshot": {
                "total": 225,
                "remainingAtAcceptance": 123,
                "remainingWithPrimaryOwnerAtAcceptance": 123,
                "unownedRemaining": 0,
            },
        },
        "workstreams": [
            {"id": "TECHNOLOGY", "primaryExecutive": "imani", "primaryExecutiveName": "Imani Heartbeat", "escalationPath": "Imani → Michelle → Founder if reserved/blocked", "remainingLaunchWork": "YES"},
            {"id": "OPERATIONS", "primaryExecutive": "michelle", "primaryExecutiveName": "Michelle Northstar", "escalationPath": "Michelle coordinates; Founder only for reserved/blockers/acceptance", "remainingLaunchWork": "YES"},
            {"id": "MARKETING", "primaryExecutive": "nia", "primaryExecutiveName": "Nia Prism", "escalationPath": "Nia → Michelle → Founder only for reserved claims", "remainingLaunchWork": "YES"},
            {"id": "FINANCE", "primaryExecutive": "kimberly", "primaryExecutiveName": "Kimberly Walker (AI)", "escalationPath": "Kimberly/Founder reserved. Michelle does not escalate routine finance implementation.", "remainingLaunchWork": "YES"},
            {"id": "EXPERIENCE", "primaryExecutive": "nia", "primaryExecutiveName": "Nia Prism", "escalationPath": "Nia → Michelle; Imani implements technical corrections", "remainingLaunchWork": "YES"},
            {"id": "LEARNING", "primaryExecutive": "nia", "primaryExecutiveName": "Nia Prism", "escalationPath": "Nia → Michelle → Founder only for material curriculum change", "remainingLaunchWork": "YES"},
            {"id": "COMMUNITY", "primaryExecutive": "nia", "primaryExecutiveName": "Nia Prism", "escalationPath": "Nia experience; Michelle operations; Imani platform/security", "remainingLaunchWork": "YES"},
            {"id": "INNOVATION", "primaryExecutive": "nia", "primaryExecutiveName": "Nia Prism", "escalationPath": "Nia → Michelle. Use August Launch Row 6; do not create a second deferred-enhancement register.", "remainingLaunchWork": "YES"},
            {"id": "LEGAL", "primaryExecutive": "imani", "primaryExecutiveName": "Imani Heartbeat", "humanLegalBoundary": "Imani owns operational/risk implementation only. Legal conclusions require a qualified human legal expert. Legal signature/Founder-reserved acceptance: Kimberly Walker. Michelle coordinates/escalates.", "escalationPath": "Imani identifies risk → Michelle escalates → human legal expert for judgment → Founder for signature/acceptance", "remainingLaunchWork": "YES"},
        ],
        "statusIntegrity": {
            "row15FounderAcceptance": "accepted",
            "row16FounderAcceptance": "accepted",
            "row17FounderAcceptance": "accepted",
            "row18FounderAcceptance": "accepted",
            "row19FounderAcceptance": "accepted",
            "workbookStatusesPreservedExceptRow18": True,
        },
        "formerPerfect10CurrentOwnershipReferences": 0,
        "potentialDuplicates": DUPLICATES,
        "criticalPath": CRITICAL_PATH,
        "criticalPathNote": "Derived from the full remaining August Launch plan in list order, with Journey completion (133–135), legal implementation/human review (32, 34), testing/regression, Go/No-Go, freeze, and launch-day activation on the spine. Banking (10–11) is on-path because settlement/reconciliation depend on it. Launch date is not changed.",
        "parallelExecution": {
            "imani": ["Tech remaining: monitoring/DR/vendors/credentials/AI controls", "Journey implementation for 133–135 after Nia specifies", "Email/analytics/security/testing spine"],
            "nia": ["Founder media 49–51, 121, 137–139, 142", "Journey specification/verification 133–135, 196–198", "Marketing 76–88, 199–200, 210, 221"],
            "michelle": ["Workstreams/authority/access coordination 18–21", "Support operations 153–157", "Testing triage and readiness 188, 191, 202–216, 222–225"],
            "kimberly": ["Banking/licenses/tax/insurance 10–14, 166", "Finance ops 163–165, 172–173", "Go/No-Go 217 and Founder announcement approval 221"],
        },
        "scheduleLaunchThreat": "YES — 125 remaining rows include Journey completion, testing, legal implementation, banking, and Go/No-Go before August 19. Launch date is not changed in Row 18.",
        "founderActionQueue": [
            {"id": f"FA-{r['spreadsheetRow']}", "row": str(r["spreadsheetRow"]), "decision": r["deliverable"], "whyFounderRequired": "Founder-reserved identity, filing, signature, policy, Go/No-Go, or Founder announcement.", "deadlineDependency": r["originalDependency"], "recommendation": r["nextAction"], "impactIfDelayed": "Blocks downstream remaining launch work on the critical path or prevents lawful/operational launch."}
            for r in remaining if r["founderActionRequired"]
        ],
        "humanExpertQueue": [
            {"id": f"HE-{r['spreadsheetRow']}", "row": str(r["spreadsheetRow"]), "action": r["deliverable"], "whyHumanExpert": "Qualified human professional required for legal, tax, or privacy conclusions. Agents implement approved text only.", "note": r["nextAction"]}
            for r in remaining if r["humanExpertRequired"]
        ],
        "flaggedNotAdded": [],
        "executiveQueueCounts": {
            "michelle": owners.get("michelle", 0),
            "imani": owners.get("imani", 0),
            "nia": owners.get("nia", 0),
            "kimberly": owners.get("kimberly", 0),
        },
        "rows": records,
        "remainingLaunchCritical": remaining,
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(adapter, ensure_ascii=False, indent=2), encoding="utf-8")
    print("SOURCE", source)
    print("TOTAL", len(records))
    print("STATUS", dict(statuses))
    print("REMAINING", len(remaining), "UNOWNED", 0)
    print("QUEUES", dict(owners))
    print("FOUNDER_Q", len(adapter["founderActionQueue"]), "HUMAN_Q", len(adapter["humanExpertQueue"]))
    print("WROTE", OUT)


if __name__ == "__main__":
    main()
